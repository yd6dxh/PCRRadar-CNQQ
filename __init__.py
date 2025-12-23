from json import load, dump, loads
from nonebot import get_bot, on_command
import hoshino
from hoshino import priv
from hoshino.typing import NoticeSession, CQEvent
from hoshino.modules.priconne import chara
from hoshino.modules.priconne import _pcr_data
from hoshino.util import DailyNumberLimiter, FreqLimiter
from .pcrclientBL import pcrclient, ApiException, bsdkclient
import asyncio
from asyncio import Lock
from os.path import dirname, join, exists
from copy import deepcopy
from traceback import format_exc
from .safeservice import SafeService
from datetime import datetime, timedelta
import time
import pytz
import openpyxl
from openpyxl.styles import Font
from .excel_handle import excel_data, sort_excel_with_styles, get_cell_size, Initialized_Data
from .SY_handle import has_claimed_reward, get_KRANK, SY_data
from .geetest import public_address

# --- 核心网络库导入 ---
try:
    from .aiorequests import get as async_get
except ImportError:
    async_get = None

sv = SafeService('深域查询')

curpath = dirname(__file__)
config = join(curpath, 'binds.json')
history = join(curpath, 'history.json')
root = {
    'arena_bind' : {},
    'config':{},
    'clan_bind':{}
}

_flmt = FreqLimiter(60)
cache = {}
client = None
lck = Lock()

if exists(config):
    with open(config) as fp:
        root = load(fp)

binds = root['arena_bind']
cfg = root['config']
clan_binds = root['clan_bind']
if exists(history):
    with open(history) as hi:
        root2 = load(hi)
else:
    root2 = {'history': {}}
clan_history = root2['history']

# --- 全局变量初始化 ---
captcha_lck = Lock()
with open(join(curpath, 'account.json')) as fp:
    acinfo = load(fp)

bot = get_bot()
validate = None
validating = False
acfirst = False
sss = 1  # 修复：初始化 sss 变量，防止 query 函数报错

# --- 自动过码逻辑 ---
async def captchaVerifierV2(gt, challenge, userid):
    global validating, async_get
    validating = True
    captcha_cnt = 0
    
    if not async_get:
        sv.logger.error("未找到 aiorequests.get，无法自动过码，尝试手动")
        return await captchaVerifier(gt, challenge, userid)

    while captcha_cnt < 5:
        captcha_cnt += 1
        try:
            sv.logger.info(f'正在尝试自动过码 (第{captcha_cnt}次)...')
            await asyncio.sleep(1)
            url = f"https://pcrd.tencentbot.top/geetest_renew?captcha_type=1&challenge={challenge}&gt={gt}&userid={userid}&gs=1"
            header = {"Content-Type": "application/json", "User-Agent": "pcrjjc2/1.0.0"}

            response = await async_get(url=url, headers=header)
            res = loads(await response.content)
            
            if "uuid" not in res: continue
            uuid = res["uuid"]
            
            ccnt = 0
            while ccnt < 10:
                ccnt += 1
                await asyncio.sleep(5)
                check_res = loads(await (await async_get(url=f"https://pcrd.tencentbot.top/check/{uuid}", headers=header)).content)
                
                if "queue_num" in check_res:
                    await asyncio.sleep(min(int(check_res["queue_num"]), 3) * 10)
                else:
                    info = check_res.get("info")
                    if info in ["fail", "url invalid"]: break
                    elif info == "in running": continue
                    elif isinstance(info, dict) and 'validate' in info:
                        sv.logger.info("自动过码成功！")
                        validating = False
                        return info["challenge"], info["gt_user_id"], info["validate"]
        except Exception as e:
            sv.logger.error(f"过码异常: {e}")

    await sendToAdmin('自动过码失败，已切换为手动模式。')
    v = await captchaVerifier(gt, challenge, userid)
    validating = False
    return challenge, userid, v

# --- 手动过码逻辑 (已修复死锁) ---
async def captchaVerifier(gt, challenge, userid):
    global validate
    # 如果锁已被占用，先强制释放
    if captcha_lck.locked():
        try: captcha_lck.release()
        except: pass

    url = f"{public_address}/geetest?captcha_type=1&challenge={challenge}&gt={gt}&userid={userid}&gs=1"
    await bot.send_private_msg(user_id=acinfo['admin'], message=f'请完成验证：\n{url}\n完成后回复：/pcrvalx [内容]')
    
    await captcha_lck.acquire() # 阻塞，直到 /pcrvalx 释放锁
    return validate

# --- 消息处理 ---
async def sendToAdmin(msg):
    await bot.send_private_msg(user_id=acinfo['admin'], message=msg)

async def errlogger(msg):
    sv.logger.error(f"登录错误: {msg}")
    await sendToAdmin(f'pcrjjc2登录错误：{msg}')

bclient = bsdkclient(acinfo, captchaVerifierV2, errlogger)
client = pcrclient(bclient)
qlck = Lock()

# 登录只做一次：避免 query/query2/query3/query4 重复 login
async def ensure_login():
    global sss
    if sss == 1:
        await client.login()
        sss = 0

async def ensure_login():
    """只在需要时登录一次（由 sss 控制）"""
    global sss
    if sss == 1:
        await client.login()
        sss = 0

async def callapi_with_relogin(path: str, data: dict, retries: int = 1):
    """
    调用 API：失败则重登并重试（解决超时/断线后无法恢复的问题）
    retries=1 表示：失败后最多再试 1 次
    """
    global sss
    async with qlck:
        for attempt in range(retries + 1):
            try:
                await ensure_login()
                return await client.callapi(path, data)
            except Exception:
                if attempt >= retries:
                    raise
                # 关键：认为登录已失效，强制下次重新 login
                sss = 1
                # 可选：给一点缓冲，避免立刻重登撞到网络抖动
                await asyncio.sleep(1)

@on_command('/pcrvalx')
async def pcr_manual_val(session):
    global validate
    if session.ctx['user_id'] == acinfo['admin']:
        validate = session.ctx['message'].extract_plain_text().replace('/pcrvalx', '').strip()
        if captcha_lck.locked():
            captcha_lck.release()
            await session.send("验证已提交")

# --- 核心查询函数 (已修复 sss 引用) ---

async def query(id: str):
    res = await callapi_with_relogin(
        '/profile/get_profile',
        {'target_viewer_id': int(id)},
        retries=1
    )
    return res['user_info']

async def query2(id: str):
    return await callapi_with_relogin(
        '/profile/get_profile',
        {'target_viewer_id': int(id)},
        retries=1
    )

async def query3(name):
    return await callapi_with_relogin(
        '/clan/search_clan',
        {
            'clan_name': str(name),
            "join_condition": 1,
            "member_condition_range": 0,
            "activity": 0,
            "clan_battle_mode": 0
        },
        retries=1
    )

async def query4(clan_id):
    return await callapi_with_relogin(
        '/clan/others_info',
        {'clan_id': int(clan_id)},
        retries=1
    )

# --- 指令部分 (修复 IndexError) ---

@sv.on_prefix(['国服绑定'])
async def pcr_bind_fixed(bot, ev: CQEvent):
    args = ev.message.extract_plain_text().split()
    gid = str(ev.group_id)
    if not args:
        await bot.finish(ev, '用法：国服绑定 [游戏ID] [@某人]', at_sender=True)
    
    ID = args[0]
    target_uid = str(ev.user_id)
    for seg in ev.message:
        if seg.type == 'at':
            target_uid = str(seg.data['qq'])
            break

    if gid not in binds:
        binds[gid] = {}
        cfg[gid] = {'admin': None, 'time' : 23}
    
    try:
        res = (await query2(ID))['user_info']
        binds[gid][ID] = {'id': str(ID), 'uid': target_uid, 'gid': gid, 'bindtype': '1'}
        save_binds()
        await bot.finish(ev, f'[{res["user_name"]}] 绑定成功！')
    except Exception as e:
        await bot.finish(ev, f'绑定失败：{e}')

# --- 辅助函数 ---
def save_binds():
    with open(config, 'w') as fp: dump(root, fp, indent=4)
def save_history():
    with open(history, 'w') as hi: dump(root2, hi, indent=4)

@sv.on_prefix('国服今日登录状态')
async def on_query_today_login_status(bot, ev):
    global binds, lck
    uid = str(ev.user_id)
    gid = str(ev.group_id)
    
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    key = f"{gid}"
    if not _flmt.check(key) and uid not in hoshino.config.SUPERUSERS:
        await bot.send(ev, f'操作太频繁，请在{int(_flmt.left_time(key))}秒后再试')
        return
    _flmt.start_cd(key)
    async with lck:
        if not gid in binds:
            await bot.finish(ev, f'名单没人！', at_sender=True)
        try:
            st = ""
            await bot.send(ev, "正在获取数据，请耐心等待")
            for uid, data in binds[gid].items():
                try:
                    res = await query(uid)
                    timeStamp = res["last_login_time"]
                    timeArray = time.localtime(timeStamp)
                    otherStyleTime = time.strftime("%Y--%m--%d %H:%M:%S", timeArray)
                    if has_claimed_reward(timeStamp):
                        login = '今日已登录'
                    else:
                        login = '今日未登录'
                    st = st + f'''[{uid}]昵称：{res["user_name"]} {login} {otherStyleTime}\n'''
                except ApiException as e:
                    await bot.send(ev, f'ID{uid}查询出错，{e}', at_sender=True)
            await bot.send(ev, st)
        except ApiException as e:
            await bot.send(ev, f'error，{e}', at_sender=True)

@sv.on_prefix('国服今日未登录名单')
async def on_query_today_not_login_list(bot, ev):
    global binds, lck
    uid = str(ev.user_id)
    gid = str(ev.group_id)
    
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    key = f"{gid}"
    if not _flmt.check(key) and uid not in hoshino.config.SUPERUSERS:
        await bot.send(ev, f'操作太频繁，请在{int(_flmt.left_time(key))}秒后再试')
        return
    _flmt.start_cd(key)
    async with lck:
        if not gid in binds:
            await bot.finish(ev, f'名单没人！', at_sender=True)
        try:
            await bot.send(ev, "正在获取数据，请耐心等待")
            n = 0
            st = ""
            for uid, data in binds[gid].items():
                try:
                    res = await query(uid)
                    timeStamp = res["last_login_time"]
                    timeArray = time.localtime(timeStamp)
                    otherStyleTime = time.strftime("%Y--%m--%d %H:%M:%S", timeArray)
                    if has_claimed_reward(timeStamp):
                        login = '今日已登录'
                    else:
                        login = '今日未登录'
                        n += 1
                        st = st + f'''[{uid}]昵称：{res["user_name"]} {login} {otherStyleTime}\n'''
                except ApiException as e:
                    await bot.send(ev, f'ID{uid}查询出错，{e}', at_sender=True)
            if n== 0:
                await bot.send(ev, "名单中今日没有未登录的玩家")
            else:
                await bot.send(ev, st)
        except ApiException as e:
            await bot.send(ev, f'error，{e}', at_sender=True)

@sv.on_rex(r'^(国服名单|国服监控名单)$')
async def send_arena_sub_status(bot,ev):
    global binds, lck
    uid = str(ev.user_id)
    gid = str(ev.group_id)

    
    if not gid in binds:
        await bot.finish(ev, f'名单没人！', at_sender=True)
        return
    else:
        st = ""
        for uid, data in binds[gid].items():
            QID = data["uid"]
            st = st + f'''游戏ID：{uid} 通知QQ：{QID}\n'''
        await bot.finish(ev, st)

@sv.on_prefix(['删除国服绑定'])
async def delete_arena_sub(bot,ev):
    global binds, lck
    args = ev.message.extract_plain_text().split()
    uid = str(ev.user_id)
    gid = str(ev.group_id)
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    ID = str(args[0])
    if not gid in binds:
        await bot.finish(ev, f'名单没人！', at_sender=True)

    async with lck:
        try:
            del binds[gid][ID]
            save_binds()
        except:
            await bot.finish(ev, f'不存在的绑定！', at_sender=True)

    await bot.finish(ev, f'已移除{ID}', at_sender=True)

@sv.on_prefix(['国服推送时间设定'])
async def times(bot, ev: CQEvent):
    args = ev.message.extract_plain_text().split()
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    if not args or len(args)>1:
        await bot.finish(ev, '请输入 国服推送时间设定+小时 中间用空格隔开。', at_sender=True)
    if not gid in binds:
        binds[gid] = {}
        cfg[gid] = {
            'admin': None,
            'time' : 23,
        }

    try:
        time = int(args[0])
    except:
        await bot.finish(ev, '请输入 国服绑定+ID+QQ号（可忽略） 中间用空格隔开。', at_sender=True)
    last = cfg.get(gid) or {}

    cfg[gid] = {
            'admin': last.get('admin'),
            'time': time
            }
    save_binds()
    await bot.finish(ev, f'设定成功！')

@sv.on_prefix(['国服会长设定'])
async def set_guild_admin(bot, ev: CQEvent):
    args = ev.message.extract_plain_text().split()
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    at = 0
    for seg in ev.message:
        if seg.type == 'at' and seg.data['qq'] != 'all':
            uid = int(seg.data['qq'])
            at = 1
            break
    if at != 1:
        try:
            uid = int(args[0])
        except:
            await bot.finish(ev, '请输入 国服会长设定+QQ号（或@）  中间用空格隔开。', at_sender=True)
        if not args or len(args)>1:
            await bot.finish(ev, '请输入 国服会长设定+QQ号（或@） 中间用空格隔开。', at_sender=True)

    if not gid in binds:
        binds[gid] = {}
        cfg[gid] = {
            'admin': None,
            'time' : 23,
        }
    last = cfg.get(gid) or {}

    cfg[gid] = {
            'admin': uid,
            'time': last.get('time', 23),
            }
    save_binds()
    await bot.finish(ev, f'设定成功！')

@sv.on_prefix(['国服绑定公会'])
async def bind_clan(bot, ev: CQEvent):
    args = ev.message.extract_plain_text().split()
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    if not args:
        await bot.finish(ev, '请输入 国服绑定公会+公会名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
    name = args[0]
    clan_id = 0
    try:
        res = await query3(name)
        if len(res["list"]) > 1:
            if len(args) < 2:
                msg = ''
                for i in range(0, len(res["list"])):
                    msg += f'会长名：{res["list"][i]["leader_name"]}\n'
                await bot.send(ev,msg)
                await bot.finish(ev, '存在重复名公会，请使用国服绑定公会+公会名+会长名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
            leader_name = args[1]
            for i in range(0, len(res["list"])):
                if leader_name == res["list"][i]["leader_name"]:
                    clan_id = res["list"][i]["clan_id"]
            if clan_id == 0:
                await bot.finish(ev, '存在重复名公会，请使用国服绑定公会+公会名+会长名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
        else:
            clan_id = res["list"][0]["clan_id"]
        clan_name = res["list"][0]["clan_name"]
    except:
        await bot.finish(ev, '获取公会信息失败，请确认公会名是否正确、公会是否可搜索!', at_sender=True)
        
    if not gid in binds:
        binds[gid] = {}
        cfg[gid] = {
            'admin': None,
            'time' : 23,
        }
        
    clan_binds[gid] = {
                'clan_id': str(clan_id),
                'clan_name':str(clan_name),
            }
    clan_history[clan_id]={
                'clan_id': str(clan_id),
                'clan_name':str(clan_name),
    }
    save_binds()
    save_history()
    await bot.finish(ev, f'绑定成功！')

@sv.on_fullmatch(['国服更新公会信息'])
async def update_clan_info(bot, ev: CQEvent):
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    if not gid in clan_binds:
        await bot.finish(ev, '请输入 国服绑定公会+公会名（必须准确） 中间用空格隔开，以绑定工会', at_sender=True)
    if not gid in binds:
        binds[gid] = {}
        cfg[gid] = {
            'admin': None,
            'time' : 23,
        }
    clan_id = clan_binds[gid]['clan_id']
    res2 = await query4(clan_id)
    members = res2["clan"]["members"]
# 提取所需字段
    result = [
    {
        "viewer_id": member["viewer_id"],
        "name": member["name"],
        "last_login_time": member["last_login_time"]
    }
    for member in members
    ]
    result_viewers = {str(user['viewer_id']) for user in result}
    msg = '本次导入了以下用户：\n'
    addnum = 0
    for data in result:
        id = str(data['viewer_id'])
        if id in binds[gid]:
            continue
        
        binds[gid][id] = {
                'id': str(id),
                'uid': uid,
                'gid': gid,
                'bindtype': '0'
            }
        msg += f'{data["name"]}({id})\n'
        addnum = 1
    if addnum == 0:
        msg += '本次未新增用户~\n'
    msg += '本次删除了以下用户：\n'
    delnum = 0
    for viewer_id in list(binds.get(gid, {}).keys()):
        if viewer_id not in result_viewers and binds[gid][viewer_id].get('bindtype', 0) != '1':
            del binds[gid][viewer_id]
            msg += f'ID：{viewer_id}\n'
            delnum +=1
    if delnum == 0:
        msg += '本次未删除用户~'
    save_binds()
    
    await bot.finish(ev, f'{msg}\n请注意，导入的ID，QQ号均默认为消息发送人，如有需要可以替换绑定')

@sv.on_fullmatch(['国服清理导入数据'])    
async def cleanup_imported_data(bot, ev: CQEvent):
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    to_remove = [
        viewer_id
        for viewer_id, data in binds.get(gid, {}).items()
        if data.get('bindtype', '0') == '0'
    ]
    print(to_remove)
    for viewer_id in to_remove:
        del binds[gid][viewer_id]
    save_binds()
    await bot.finish(ev, f'清理已完成')

@sv.on_fullmatch('国服生成深域表')
async def gen_sy_table_default(bot,ev):
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    sheet = Initialized_Data()
    numx = 1
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    key = f"{gid}"
    if not _flmt.check(key) and uid not in hoshino.config.SUPERUSERS:
        await bot.send(ev, f'操作太频繁，请在{int(_flmt.left_time(key))}秒后再试')
        return
    if not gid in binds:
            await bot.finish(ev, f'名单没人！', at_sender=True)
    if len(binds[gid]) >= 36:
        await bot.finish(ev, f"{gid} 下的 uid 数量超出上限！")
    _flmt.start_cd(key)
    async with lck:
        try:
            await bot.send(ev, "正在获取数据，请耐心等待")
            inclan = 0
            if gid in clan_binds:
                clan_id = clan_binds[gid]['clan_id']
                res2 = await query4(clan_id)
                members = res2["clan"]["members"]
                # 提取所需字段
                result = [
                {
                "viewer_id": member["viewer_id"],
                "name": member["name"],
                "last_login_time": member["last_login_time"]
                }
                for member in members
                ]
                result_viewers = {str(user['viewer_id']) for user in result}
                inclan = 1
            for uid, data in binds[gid].items():
                try:
                    i = 0
                    res5 = await query2(uid)
                    res4 = res5['user_info']
                    numx = numx + 1
                    SY_data(res4,res5,sheet,uid,data,numx)
                    '''
                    for i in range(0, 5):
                        LIE = chr(74 + i)
                        if int(res5['quest_info']['talent_quest'][i]['clear_count']) < 59 :
                            if int(res5['quest_info']['talent_quest'][i]['clear_count']) < 51:
                                cell = sheet[f'{LIE}{numx}']
                                cell.font = Font(color="FF0000")
                            else:
                                cell = sheet[f'{LIE}{numx}']
                                cell.font = Font(color="800080")
                    '''
                    if inclan:
                        if str(uid) not in result_viewers:
                            cell = sheet[f'C{numx}']
                            cell.font = Font(color="FF0000")
                except ApiException as e:
                    await bot.send(ev, f'ID{uid}查询出错，{e}', at_sender=True)
            data = list(sheet.iter_rows(min_row=2, values_only=True))

            # 根据 G 列（第7列）进行降序排序
            sort_excel_with_styles(sheet, sort_column_index=6)
            base64_str = excel_data(sheet,'0')
            await bot.finish(ev, f"[CQ:image,file={base64_str}]")

        except ApiException as e:
            await bot.send(ev, f'error，{e}', at_sender=True)

@sv.on_prefix('国服生成公会深域表')
async def gen_sy_table_by_clan_name(bot,ev):
    args = ev.message.extract_plain_text().split()
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    sheet = Initialized_Data()
    numx = 1
    if gid == '901952944' or gid == 901952944:
        key = f"{uid}"
    else:
        u_priv = priv.get_user_priv(ev)
        if u_priv < sv.manage_priv:
            await bot.finish(ev, '权限不足', at_sender=True)
        key = f"{gid}"
    if not _flmt.check(key) and uid not in hoshino.config.SUPERUSERS:
        await bot.send(ev, f'操作太频繁，请在{int(_flmt.left_time(key))}秒后再试')
        return
    if not args:
        await bot.finish(ev, '请使用国服生成公会深域表+公会名+会长名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
    name = args[0]
    clan_id = 0
    try:
        res = await query3(name)
        if len(res["list"]) > 1:
            if len(args) < 2:
                msg = ''
                for i in range(0, len(res["list"])):
                    msg += f'会长名：{res["list"][i]["leader_name"]}  {res["list"][i]["member_num"]}人\n'
                await bot.send(ev,msg)
                await bot.finish(ev, '存在重复名公会，请使用国服生成公会深域表+公会名+会长名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
            leader_name = args[1]
            for i in range(0, len(res["list"])):
                if leader_name == res["list"][i]["leader_name"]:
                    clan_id = res["list"][i]["clan_id"]
            if clan_id == 0:
                await bot.finish(ev, '存在重复名公会，请使用国服生成公会深域表+公会名+会长名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
        else:
            clan_id = res["list"][0]["clan_id"]
        clan_name = res["list"][0]["clan_name"]
    except:
        await bot.finish(ev, '获取公会信息失败，请确认公会名是否正确、公会是否可搜索!', at_sender=True)
    if clan_id != 0:
        clan_history[clan_id]={
                'clan_id': str(clan_id),
                'clan_name':str(clan_name),
            }
        save_history()
        _flmt.start_cd(key)
        async with lck:
            try:
                await bot.send(ev, "正在获取数据，请耐心等待")
                res2 = await query4(clan_id)
                members = res2["clan"]["members"]
                    # 提取所需字段
                result = [
                {
                "viewer_id": member["viewer_id"],
                "name": member["name"],
                "last_login_time": member["last_login_time"]
                }
                for member in members
                ]
                result_viewers = {str(user['viewer_id']) for user in result}
                for uid in result_viewers:
                    try:
                        data = '0'
                        i = 0
                        res5 = await query2(uid)
                        res4 = res5['user_info']
                        numx = numx + 1
                        SY_data(res4,res5,sheet,uid,data,numx)
                        '''
                        for i in range(0, 5):
                            LIE = chr(74 + i)
                            if int(res5['quest_info']['talent_quest'][i]['clear_count']) < 59 :
                                if int(res5['quest_info']['talent_quest'][i]['clear_count']) < 51:
                                    cell = sheet[f'{LIE}{numx}']
                                    cell.font = Font(color="FF0000")
                                else:
                                    cell = sheet[f'{LIE}{numx}']
                                    cell.font = Font(color="800080")
                        '''
                    except ApiException as e:
                        await bot.send(ev, f'ID{uid}查询出错，{e}', at_sender=True)
                sheet.delete_cols(2)
                data = list(sheet.iter_rows(min_row=2, values_only=True))
                
            # 根据 G 列（第6列）进行降序排序
                sort_excel_with_styles(sheet, sort_column_index=5)
                base64_str = excel_data(sheet,f'{clan_name}')
                await bot.finish(ev, f"[CQ:image,file={base64_str}]")

            except ApiException as e:
                await bot.send(ev, f'error，{e}', at_sender=True)

@sv.on_prefix('国服生成ID深域表')
async def gen_sy_table_by_clan_id(bot,ev):
    args = ev.message.extract_plain_text().split()
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    sheet = Initialized_Data()
    numx = 1
    if not priv.check_priv(ev, priv.SUPERUSER):
        await bot.finish(ev, '无权进行该操作！', at_sender=True)
    clan_id = args[0]
    if not args:
        await bot.finish(ev, '请使用国服生成公会深域表+公会名+会长名（必须准确） 中间用空格隔开。公会必须处于可搜索状态', at_sender=True)
    else:
        async with lck:
            try:
                await bot.send(ev, "正在获取数据，请耐心等待")
                res2 = await query4(clan_id)
                members = res2["clan"]["members"]
                clan_name = res2['clan']['detail']["clan_name"]
                    # 提取所需字段
                result = [
                {
                "viewer_id": member["viewer_id"],
                "name": member["name"],
                "last_login_time": member["last_login_time"]
                }
                for member in members
                ]
                result_viewers = {str(user['viewer_id']) for user in result}
                for uid in result_viewers:
                    try:
                        data = '0'
                        i = 0
                        res5 = await query2(uid)
                        res4 = res5['user_info']
                        numx = numx + 1
                        SY_data(res4,res5,sheet,uid,data,numx)
                        for i in range(0, 5):
                            LIE = chr(74 + i)
                            if int(res5['quest_info']['talent_quest'][i]['clear_count']) < 59 :
                                if int(res5['quest_info']['talent_quest'][i]['clear_count']) < 51:
                                    cell = sheet[f'{LIE}{numx}']
                                    cell.font = Font(color="FF0000")
                                else:
                                    cell = sheet[f'{LIE}{numx}']
                                    cell.font = Font(color="800080")
                    except ApiException as e:
                        await bot.send(ev, f'ID{uid}查询出错，{e}', at_sender=True)
                sheet.delete_cols(2)
                data = list(sheet.iter_rows(min_row=2, values_only=True))
                
            # 根据 G 列（第6列）进行降序排序
                sort_excel_with_styles(sheet, sort_column_index=5)
                base64_str = excel_data(sheet,f'{clan_name}')
                await bot.finish(ev, f"[CQ:image,file={base64_str}]")

            except ApiException as e:
                await bot.send(ev, f'error，{e}', at_sender=True)

@sv.on_prefix('国服生成定制深域表')
async def gen_sy_table_custom_threshold(bot,ev):
    args = ev.message.extract_plain_text().split()
    u_priv = priv.get_user_priv(ev)
    if u_priv < sv.manage_priv:
        await bot.finish(ev, '权限不足', at_sender=True)
    data_dict = {}
    if len(args) != 5:
        await bot.finish(ev, '数据填写不正确，请按火 水 风 光 暗的顺序填写，不要带杠（7-9填写79，7-10填写80）', at_sender=True)
    try:
        int(args[0])
    except:
        await bot.finish(ev, '数据填写不正确，请按火 水 风 光 暗的顺序填写，不要带杠（7-9填写79，7-10填写80）', at_sender=True)
    for i in range(0, 5):  # i 变化从 1 到 5
        data_dict[i] = {
            "value": f"{args[i]}",  # 这里换成你的实际数据
        }
    gid = str(ev.group_id)
    uid = str(ev.user_id)
    sheet = Initialized_Data()
    numx = 1

    key = f"{gid}"
    if not _flmt.check(key) and uid not in hoshino.config.SUPERUSERS:
        await bot.send(ev, f'操作太频繁，请在{int(_flmt.left_time(key))}秒后再试')
        return
    if len(binds[gid]) >= 36:
        await bot.finish(ev, f"{gid} 下的 uid 数量超出上限！")
    _flmt.start_cd(key)
    async with lck:
        if not gid in binds:
            await bot.finish(ev, f'名单没人！', at_sender=True)
        try:
            await bot.send(ev, "正在获取数据，请耐心等待")
            inclan = 0
            if gid in clan_binds:
                clan_id = clan_binds[gid]['clan_id']
                res2 = await query4(clan_id)
                members = res2["clan"]["members"]
# 提取所需字段
                result = [
                {
                "viewer_id": member["viewer_id"],
                "name": member["name"],
                "last_login_time": member["last_login_time"]
                }
                for member in members
                ]
                result_viewers = {str(user['viewer_id']) for user in result}
                inclan = 1
            for uid, data in binds[gid].items():
                try:
                    i = 0
                    res5 = await query2(uid)
                    res4 = res5['user_info']
                    numx = numx + 1
                    SY_data(res4,res5,sheet,uid,data,numx)
                    for i in range(0, 5):
                        LIE = chr(74 + i)
                        if int(res5['quest_info']['talent_quest'][i]['clear_count']) < int(args[i])-10 :
                            cell = sheet[f'{LIE}{numx}']
                            cell.font = Font(color="FF0000")
                    if inclan:
                        if str(uid) not in result_viewers:
                            cell = sheet[f'C{numx}']
                            cell.font = Font(color="FF0000")
                except ApiException as e:
                    await bot.send(ev, f'ID{uid}查询出错，{e}', at_sender=True)
            data = list(sheet.iter_rows(min_row=2, values_only=True))

            # 根据 G 列（第7列）进行降序排序
            sort_excel_with_styles(sheet, sort_column_index=6)
            base64_str = excel_data(sheet,'0')
            await bot.finish(ev, f"[CQ:image,file={base64_str}]")

        except ApiException as e:
            await bot.send(ev, f'error，{e}', at_sender=True)


@sv.scheduled_job('interval', hours=1)
async def on_arena_schedule():
    global binds, cfg, lck
    bot = get_bot()
    tz = pytz.timezone('Asia/Shanghai')

    # 保活/容错：避免接口偶发失败影响主流程
    try:
        await query4(1)
    except:
        pass

    async with lck:
        bind_cache = deepcopy(binds)
        cfg_cache = deepcopy(cfg)

    now_hour = datetime.now(tz).hour

    for gid, uid_data in bind_cache.items():
        gid_str = str(gid)
        group_cfg = cfg_cache.get(gid_str, {})
        push_hour = group_cfg.get("time", 23)
        if push_hour != now_hour:
            continue

        admin_id = group_cfg.get("admin")
        header = f"[CQ:at,qq={admin_id}]\n" if admin_id else ""

        lines = []
        n_not_login = 0

        for game_uid, data in uid_data.items():
            try:
                res = await query(game_uid)
                timeStamp = res["last_login_time"]
                otherStyleTime = time.strftime("%Y--%m--%d %H:%M:%S", time.localtime(timeStamp))

                # 与“国服今日登录状态/未登录名单”保持一致：has_claimed_reward 为真则视作今日已登录
                if has_claimed_reward(timeStamp):
                    continue

                n_not_login += 1
                qq_to_at = data.get("uid")
                lines.append(
                    f'[CQ:at,qq={qq_to_at}][{game_uid}]昵称：{res["user_name"]} 今日未登录 {otherStyleTime}'
                )
            except ApiException as e:
                await bot.send_group_msg(group_id=int(gid), message=f'ID{game_uid}查询出错，{e}')
            except Exception as e:
                await bot.send_group_msg(group_id=int(gid), message=f'ID{game_uid}查询异常，{e}')

        if n_not_login == 0:
            msg = header + "今日登记的用户均已登录"
        else:
            msg = header + "\n".join(lines)

        await bot.send_group_msg(group_id=int(gid), message=msg)
