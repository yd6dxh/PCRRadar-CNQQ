import io
import base64
import copy
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
from openpyxl.styles import Border, Side,Font
from PIL import Image, ImageDraw, ImageFont
import openpyxl

font_path = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
font = ImageFont.truetype(font_path, 24)
def Initialized_Data():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet[f'A{1}'] = 'ID'
    sheet[f'B{1}'] = 'QQ'
    sheet[f'C{1}'] = '昵称'
    sheet[f'D{1}'] = '等级'
    sheet[f'E{1}'] = '战斗力'
    sheet[f'F{1}'] = '角色数'
    sheet[f'G{1}'] = '深域等级'
    sheet[f'H{1}'] = '离升级需要经验'
    sheet[f'I{1}'] = '最后一次登入时间'
    sheet[f'J{1}'] = '深域火通关'
    sheet[f'K{1}'] = '深域水通关'
    sheet[f'L{1}'] = '深域风通关'
    sheet[f'M{1}'] = '深域光通关'
    sheet[f'N{1}'] = '深域暗通关'
    return sheet
def excel_data(sheet,name):
            width = 3  # 手动加宽的数值 
        # 单元格列宽处理
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
                for col, value in dims.items():
                    sheet.column_dimensions[get_column_letter(col)].width = value + width
            align = Alignment(horizontal='center', vertical='center',wrapText=True)
        # 两层循环遍历所有有数据的单元格
            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    sheet.cell(i, j).alignment = align
            # 定义边框样式（细线）
            thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

            # 获取表格区域的最大行数和列数
            max_row = sheet.max_row
            max_col = sheet.max_column

            # 遍历每个单元格并设置边框
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = thin_border
    # 绘制图形
            total_width = sum(get_cell_size(sheet, chr(65 + col), 1)[0] for col in range(max_col))
            total_height = sum(get_cell_size(sheet, 'A', row)[1] for row in range(1, max_row + 1))+40
            image = Image.new("RGB", (int(total_width), int(total_height)), "white")
            draw = ImageDraw.Draw(image)
            current_y = 0
            for row_index in range(1, max_row + 1):
                current_x = 0
                for col_index in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_index, column=col_index).value or ""
                    col_letter = chr(64 + col_index)  # A, B, C...
                     # 获取字体颜色（如果有设置）
                    cell = sheet.cell(row=row_index, column=col_index)
                    font_color = cell.font.color
                    try:
            # 默认字体颜色为黑色
                        if font_color and font_color.rgb:
                            font_color = font_color.rgb[2:]  # 获取 RGB 颜色的十六进制部分
                        else:
                            font_color = "000000"  # 黑色
                    except:
                        font_color = "000000"

            # 转换颜色格式 (0xRRGGBB)
                    font_color = "#" + font_color  # #RRGGBB格式

                    cell_width, cell_height = get_cell_size(sheet, col_letter, row_index)
                    # 绘制单元格边框
                    draw.rectangle([current_x, current_y, current_x + cell_width, current_y + cell_height], outline="black", width=2)
                    # 居中绘制文本
                    text_width, text_height = draw.textsize(str(cell_value), font=font)
                    text_x = current_x + (cell_width - text_width) / 2
                    text_y = current_y + (cell_height - text_height) / 2
                    draw.text((text_x, text_y), str(cell_value), fill=font_color, font=font)
                    current_x += cell_width
                current_y += cell_height
            if name == '0':
                remark_text = "注：名字标红时，说明此玩家不在公会中(需绑定公会)；登录时间标紫时，说明今日尚未登录，标红时，说明昨天也没登录；深域进度标紫时，表明进度略微落后，标红时，表明深域进度严重落后。"
            else:
                remark_text = f"公会名：{name}   注：登录时间标紫时，说明今日尚未登录，标红时，说明昨天也没登录；深域进度标紫时，表明进度略微落后，标红时，表明深域进度严重落后。"
            img_height = cell_height * max_row + 40
            remark_y = max_row * cell_height  # 备注行的 Y 坐标
            draw.rectangle([0, remark_y, img_height, remark_y + 40], fill="white")

            # 计算备注文本的 X 位置（居中）
            text_size = draw.textbbox((0, 0), remark_text, font=font)
            text_x = 10
            text_y = remark_y + (40 - text_size[3]) / 2

    # 写入备注文本
            draw.text((text_x, text_y), remark_text, fill="black", font=font)
            buffer = io.BytesIO()
            image.save(buffer, format="PNG")
            buffer.seek(0)
            # 转换为 Base64
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            buffer.close()
            base64_str = 'base64://' + img_base64
            return base64_str

def sort_excel_with_styles(sheet, sort_column_index, start_row=2):
    #存储每一行的值和样式信息
    rows_data = []

    for row in sheet.iter_rows(min_row=start_row, values_only=False):
        row_data = []
        for cell in row:
            row_data.append({
                'value': cell.value,
                'font': copy.copy(cell.font),         #深拷贝字体样式
                'fill': copy.copy(cell.fill),         #深拷贝填充样式
                'border': copy.copy(cell.border),     #深拷贝边框样式
                'alignment': copy.copy(cell.alignment) #深拷贝对齐方式
            })
        rows_data.append(row_data)

    #按指定列进行降序排序（sort_column_index从0开始）
    rows_data.sort(key=lambda row: row[sort_column_index]['value'], reverse=True)

    #清空原始数据和样式
    for row_index in range(start_row, sheet.max_row + 1):
        for col_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = None
            cell.font = None
            cell.fill = None
            cell.border = None
            cell.alignment = None

    #将排序后的数据和样式写回到 Excel
    for row_index, row_data in enumerate(rows_data, start=start_row):
        for col_index, cell_data in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = cell_data['value']
            cell.font = cell_data['font']
            cell.fill = cell_data['fill']
            cell.border = cell_data['border']
            cell.alignment = cell_data['alignment']

def get_cell_size(ws, col_letter, row):
    col_width = (ws.column_dimensions[col_letter].width or 8) * 10 * 1.5  # 调整为10像素单位
    row_height = (ws.row_dimensions[row].height or 15) * 2.5  # 调整行高
    return col_width, row_height
